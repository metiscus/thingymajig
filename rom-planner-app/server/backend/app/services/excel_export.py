# backend/app/services/excel_export.py
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import List, Dict, Any

# Import your SQLModel models
from app.models import Project, Rate, Task, MaterialItem

def get_role_rate(rates: List[Rate], role_name: str) -> float:
    """Helper to get rate for a given role, adjusting for hours to days."""
    rate_obj = next((r for r in rates if r.role == role_name), None)
    if not rate_obj:
        return 0.0
    # Ensure rate_obj.rate is treated as a number
    return (rate_obj.rate or 0) * 8 if rate_obj.unit == 'hour' else (rate_obj.rate or 0)

def calculate_task_total_cost(task: Task, rates: List[Rate]) -> float:
    """Calculates total cost for a single task based on efforts, travel, and materials."""
    cost = 0.0
    for role, efforts in (task.efforts or {}).items(): # Use .get() or default to empty dict for safety
        cost += (efforts or 0) * get_role_rate(rates, role)
    cost += (task.travelCost or 0)
    cost += (task.materialsCost or 0)
    return cost

def calculate_task_total_days(task: Task) -> float:
    """Calculates total days for a single task from efforts."""
    return sum((task.efforts or {}).values()) if task.efforts else 0.0

def generate_project_excel(
    project: Project,
    rates: List[Rate],
    tasks: List[Task],
    material_items: List[MaterialItem]
) -> BytesIO:
    """Generates an Excel workbook with project summary, labor, rates, and materials details."""
    workbook = Workbook()

    # --- Summary Sheet ---
    summary_sheet = workbook.active
    summary_sheet.title = 'Summary'
    
    summary_sheet.column_dimensions['A'].width = 30
    summary_sheet.column_dimensions['B'].width = 40

    summary_sheet.append(['Project Name:', project.name])
    summary_sheet.cell(row=1, column=1).font = Font(bold=True)
    summary_sheet.cell(row=1, column=2).font = Font(bold=True)
    summary_sheet.append(['Project Description:', project.description])
    summary_sheet.append([]) # Empty row

    # Calculate totals
    all_roles = sorted(list(set(r.role for r in rates)))

    total_days_per_role: Dict[str, float] = {role: 0.0 for role in all_roles}
    for task in tasks:
        for role, efforts in (task.efforts or {}).items():
            if role in total_days_per_role:
                total_days_per_role[role] += (efforts or 0)

    total_cost_per_role: Dict[str, float] = {role: 0.0 for role in all_roles}
    total_labor_cost = 0.0
    for role in all_roles:
        rate = get_role_rate(rates, role)
        cost = total_days_per_role[role] * rate
        total_cost_per_role[role] = cost
        total_labor_cost += cost
    
    total_task_travel_cost = sum((t.travelCost or 0) for t in tasks)
    total_task_incidental_materials = sum((t.materialsCost or 0) for t in tasks)
    total_detailed_material_expenses = sum((item.unitPrice or 0) * (item.quantity or 0) for item in material_items)

    subtotal_before_risk = total_labor_cost + total_task_travel_cost + \
                           total_task_incidental_materials + total_detailed_material_expenses
    
    risk_percentage = project.riskPercentage or 0
    risk_amount = subtotal_before_risk * (risk_percentage / 100)
    grand_total_with_risk = subtotal_before_risk + risk_amount

    summary_sheet.append(['Total Labor Cost:', total_labor_cost])
    summary_sheet.cell(row=summary_sheet.max_row, column=2).number_format = '$#,##0.00'
    summary_sheet.append(['Total Task Travel Cost:', total_task_travel_cost])
    summary_sheet.cell(row=summary_sheet.max_row, column=2).number_format = '$#,##0.00'
    summary_sheet.append(['Total Task Incidental Materials:', total_task_incidental_materials])
    summary_sheet.cell(row=summary_sheet.max_row, column=2).number_format = '$#,##0.00'
    summary_sheet.append(['Total Detailed Material Expenses:', total_detailed_material_expenses])
    summary_sheet.cell(row=summary_sheet.max_row, column=2).number_format = '$#,##0.00'
    summary_sheet.append([]) # Empty row

    subtotal_row = summary_sheet.append(['SUBTOTAL (Before Risk):', subtotal_before_risk])
    subtotal_row[0].font = Font(bold=True)
    subtotal_row[1].font = Font(bold=True)
    subtotal_row[1].number_format = '$#,##0.00'
    
    risk_row = summary_sheet.append([f'Risk ({risk_percentage}%):', risk_amount])
    risk_row[1].number_format = '$#,##0.00'
    summary_sheet.append([]) # Empty row

    grand_total_row = summary_sheet.append(['PROJECT GRAND TOTAL (incl. Risk):', grand_total_with_risk])
    for cell in grand_total_row:
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="FF35495E", end_color="FF35495E", fill_type="solid")
        cell.font.color.rgb = 'FFFFFFFF'
    grand_total_row[1].number_format = '$#,##0.00'

    for row_idx in range(1, summary_sheet.max_row + 1):
        for col_idx in range(1, summary_sheet.max_column + 1):
            summary_sheet.cell(row=row_idx, column=col_idx).alignment = Alignment(vertical='top', horizontal='left')


    # --- Labor Details Sheet ---
    labor_sheet = workbook.create_sheet('Labor Details')
    labor_headers = ['Task Name', 'Description'] + [f'{role} (Days)' for role in all_roles] + \
                    ['Travel Cost', 'Materials Cost', 'Total Task Days', 'Total Task Cost']
    
    labor_sheet.append(labor_headers)
    for cell in labor_sheet[1]:
        cell.font = Font(bold=True)

    # Set column widths and formats
    labor_sheet.column_dimensions['A'].width = 30 # Task Name
    labor_sheet.column_dimensions['B'].width = 40 # Description
    col_idx = 3
    for _ in all_roles:
        labor_sheet.column_dimensions[get_column_letter(col_idx)].width = 15
        labor_sheet.column_dimensions[get_column_letter(col_idx)].alignment = Alignment(horizontal='center')
        col_idx += 1
    labor_sheet.column_dimensions[get_column_letter(col_idx)].width = 15 # Travel Cost
    labor_sheet.column_dimensions[get_column_letter(col_idx)].number_format = '$#,##0.00'
    col_idx += 1
    labor_sheet.column_dimensions[get_column_letter(col_idx)].width = 15 # Materials Cost
    labor_sheet.column_dimensions[get_column_letter(col_idx)].number_format = '$#,##0.00'
    col_idx += 1
    labor_sheet.column_dimensions[get_column_letter(col_idx)].width = 15 # Total Task Days
    labor_sheet.column_dimensions[get_column_letter(col_idx)].alignment = Alignment(horizontal='center')
    col_idx += 1
    labor_sheet.column_dimensions[get_column_letter(col_idx)].width = 18 # Total Task Cost
    labor_sheet.column_dimensions[get_column_letter(col_idx)].number_format = '$#,##0.00'


    for task in tasks:
        row_data = [task.name, task.description] + \
                   [((task.efforts or {}).get(role) or 0) for role in all_roles] + \
                   [task.travelCost, task.materialsCost, calculate_task_total_days(task), calculate_task_total_cost(task, rates)]
        labor_sheet.append(row_data)
        # Apply number formats to cost columns
        row = labor_sheet[labor_sheet.max_row]
        row[len(all_roles) + 2].number_format = '$#,##0.00' # Travel Cost (1-indexed column)
        row[len(all_roles) + 3].number_format = '$#,##0.00' # Materials Cost
        row[len(all_roles) + 5].number_format = '$#,##0.00' # Total Task Cost
    
    # Project Totals row
    labor_totals_row_data = ['Project Totals:', '']
    for role in all_roles:
        labor_totals_row_data.append(total_days_per_role[role])
    labor_totals_row_data.append(total_task_travel_cost)
    labor_totals_row_data.append(total_task_incidental_materials)
    labor_totals_row_data.append(sum(total_days_per_role.values()))
    labor_totals_row_data.append(total_labor_cost + total_task_travel_cost + total_task_incidental_materials)
    
    labor_sheet.append(labor_totals_row_data)
    labor_totals_row = labor_sheet[labor_sheet.max_row]
    for cell in labor_totals_row:
        cell.font = Font(bold=True)
    labor_totals_row[len(all_roles) + 2].number_format = '$#,##0.00'
    labor_totals_row[len(all_roles) + 3].number_format = '$#,##0.00'
    labor_totals_row[len(all_roles) + 5].number_format = '$#,##0.00'

    # Cost per Role row
    cost_per_role_row_data = ['Cost per Role:', '']
    for role in all_roles:
        cost_per_role_row_data.append(total_cost_per_role[role])
    labor_sheet.append(cost_per_role_row_data)
    cost_per_role_row = labor_sheet[labor_sheet.max_row]
    for i in range(2, 2 + len(all_roles)): # Start from column C (index 2 in 0-indexed list, but 3rd col in excel)
        cost_per_role_row[i].number_format = '$#,##0.00'


    # --- Rates Sheet ---
    rates_sheet = workbook.create_sheet('Rates')
    rates_sheet.append(['Role', 'Rate', 'Unit'])
    for cell in rates_sheet[1]:
        cell.font = Font(bold=True)
    rates_sheet.column_dimensions['A'].width = 25
    rates_sheet.column_dimensions['B'].width = 15
    rates_sheet.column_dimensions['C'].width = 10
    rates_sheet.column_dimensions['B'].number_format = '$#,##0.00' # Apply format to the column

    for rate in rates:
        rates_sheet.append([rate.role, rate.rate, rate.unit])


    # --- Materials Sheet ---
    materials_sheet = workbook.create_sheet('Materials')
    materials_sheet.append(['Line Item', 'Vendor', 'Category', 'Unit Price', 'Quantity', 'Subtotal', 'Comment'])
    for cell in materials_sheet[1]:
        cell.font = Font(bold=True)
    
    materials_sheet.column_dimensions['A'].width = 30
    materials_sheet.column_dimensions['B'].width = 20
    materials_sheet.column_dimensions['C'].width = 15
    materials_sheet.column_dimensions['D'].width = 15
    materials_sheet.column_dimensions['E'].width = 10
    materials_sheet.column_dimensions['F'].width = 15
    materials_sheet.column_dimensions['G'].width = 40

    materials_sheet.column_dimensions['D'].number_format = '$#,##0.00'
    materials_sheet.column_dimensions['F'].number_format = '$#,##0.00'

    for item in material_items:
        subtotal = (item.unitPrice or 0) * (item.quantity or 0)
        materials_sheet.append([
            item.lineItem, item.vendor, item.category, item.unitPrice, item.quantity, subtotal, item.comment
        ])
    
    materials_sheet.append([]) # Empty row
    total_mat_row = materials_sheet.append(['Total Detailed Material Costs:', '', '', '', '', total_detailed_material_expenses])
    total_mat_row[0].font = Font(bold=True)
    total_mat_row[5].font = Font(bold=True)
    total_mat_row[5].number_format = '$#,##0.00'

    # Save to BytesIO
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer